#!/usr/bin/env python

# IMPORTING THE LIB FOLDER
import os, sys
sys.path.insert(1,os.path.join(os.path.dirname(__file__), "..", "lib"))

from helpers import Settings

# **********************************
# ** LIBS
# ***********************************

# NUMPY (scientific computing package)
import numpy as np

# OPENPYXL (reads and writes excel files)
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.drawing import Image

from openpyxl.charts import LineChart, Reference, Series

# MATPLOTLIB (draws graphs)
import matplotlib
#print matplotlib.__version__
from matplotlib.ticker import FuncFormatter
matplotlib.use('Qt4Agg')
matplotlib.rcParams.update({'font.size': 14})
from matplotlib import pylab as p

# ORDEREDDICT
from collections import OrderedDict

from operator import add
# ***********************************


fileName    = "../../output/lfsq_egan2.xlsx"
outFileName    = "../../output/lfsq_egan2_img.xlsx"

offset      = 12

# DIMENSIONS: Sheet, Row

lines = []
lines.append({
    'color'     : '#000000',
    'row'       : 'TOTAL'});

lines.append({
    'color'     : '#0070c0',
    'row'       : 'A'});

# DIMENSIONS: Sheets, Rows
graphs = {
    'type'      : 'sheets',
    'selection' : ['AT', 'DE', 'ES']}


wb = None
ws = {}
rowNameToRowNumber = {}
timeLabels = None
startRow = 1
endRow = None
dataCol = 2

def findRowNumberByRowName(sheet = None, row = None):
    t = startRow
    while t < 10000:
        t += 1
        if isinstance(row, tuple):
            allI = True
            for i in range(len(row)):
                if row[i] != sheet.cell('%s%s'%(get_column_letter(i+1), t)).value:
                    allI = False
                    break;
            if allI == True:
                break
        else:
            #print sheet.cell('A%s'%(t)).value
            if row == sheet.cell('A%s'%(t)).value:
                break

    return t

def init(sheet = None, sheetName = None):
    global timeLabels, startRow, endRow
    #print sheet
    if sheet is None:
        if sheetName not in ws:
            ws[sheetName] = wb.get_sheet_by_name(name = sheetName)
            sheet = ws[sheetName]

    if timeLabels is None:
        timeLabels = []
        startRow = findRowNumberByRowName(sheet, "Data:") + 1
        endRow = findRowNumberByRowName(sheet, None)
        t = dataCol
        while t < 1000:
            label = sheet.cell('%s%s'%( get_column_letter(t), startRow)).value
            if label is None:
                break
            timeLabels.append(label)
            t += 1

def getData(sheet = None, sheetName = None, row = None, rowNumber = None):
    if sheet is None:
        if sheetName not in ws:
            ws[sheetName] = wb.get_sheet_by_name(name = sheetName)
            sheet = ws[sheetName]

    if rowNumber is None:
        if row not in rowNameToRowNumber:
            rowNameToRowNumber[row] = findRowNumberByRowName(sheet, row)
        rowNumber = rowNameToRowNumber[row]

    data = []

    for t in range(len(timeLabels)):
        d = sheet.cell('%s%s'%( get_column_letter(t+dataCol), rowNumber)).value
        if d == "":
            d = None
        data.append(d)

    return data

def main():
    global wb
    wb = load_workbook(filename = fileName)
    for s1 in graphs['selection']:
        dataParam = {}
        if graphs['type'] == "sheets":
            dataParam['sheet'] = s1
        elif graphs['type'] == "row":
            dataParam['row'] = s1


        p.figure(figsize=(10.0, 5.0))

        maxValue = -float("inf")
        minValue =  float("inf")

        for line in lines:
            if 'row' in line:
                s2 = line['row']
                dataParam['row'] = s2
            elif 'sheet' in line:
                s2 = line['sheet']
                dataParam['sheet'] = s2

            print dataParam
            data = getData(**dataParam)

            p.plot(range(0, len(data)), data, color=line["color"], label = s2)

        #p.xlabel(timeLabels, rotation='vertical', size = 'x-small')
        ax = p.gca()
        ax.yaxis.label.set_size('x-small')
        p.xticks(range(len(data)), timeLabels, rotation='vertical', size = 'x-small', multialignment = 'center')
        #p.xticks(map(add, range(len(data)), [-0.5] * len(data)), timeLabels, rotation='vertical', size = 'x-small', ha = 'right')

        p.legend(loc='best', prop={"size": 'small'}).draw_frame(False)
        p.grid(True)
        p.tight_layout()

        p.savefig("tttt.png")

        img = Image('tttt.png')
        img.anchor(ws[dataParam['sheet']]['B%s'%(endRow + 1)], anchortype='oneCell')
        ws[dataParam['sheet']].add_image(img)

        wb.save(outFileName)


def addGraph(sheet, c):
    global dataCol
    dataCol = c + 1
    #print dataCol
    dataParam = {}
    dataParam['sheet'] = sheet

    p.figure(figsize=(12.0, 5.0))

    init(sheet = sheet)

    #print startRow, endRow
    for i in range(startRow + 1, endRow):
        dataParam['rowNumber'] = i
        dataParam['sheet'] = sheet

        data = getData(**dataParam)

        #p.plot(range(0, len(data)), data, color = line["color"], label = s2)

        #print data
        l = []
        for j in range(1, dataCol):
            l.append(sheet.cell('%s%s'%( get_column_letter(j), i)).value)
        p.plot(range(0, len(data)), data, label = " - ".join(l))

    #p.xlabel(timeLabels, rotation='vertical', size = 'x-small')
    ax = p.gca()
    ax.yaxis.label.set_size('x-small')
    p.xticks(range(len(timeLabels)), timeLabels, rotation='vertical', size = 'x-small', multialignment = 'center')
    #p.xticks(map(add, range(len(data)), [-0.5] * len(data)), timeLabels, rotation='vertical', size = 'x-small', ha = 'right')

    p.legend(loc='best', prop={"size": 'small'}).draw_frame(False)
    p.grid(True)
    p.tight_layout()

    imgName = "_tttt" + sheet.title + ".png"

    p.savefig(imgName)
    img = Image(imgName)
    img.anchor(sheet['B%s'%(endRow + 1)], anchortype='oneCell')
    sheet.add_image(img)

    #os.remove("_tttt.png")


def addImageGraph(sheet, c):
    global dataCol
    dataCol = c + 1
    #print dataCol
    dataParam = {}
    dataParam['sheet'] = sheet

    p.figure(figsize=(12.0, 5.0))

    init(sheet = sheet)

    #print startRow, endRow
    for i in range(startRow + 1, endRow):
        dataParam['rowNumber'] = i
        dataParam['sheet'] = sheet

        data = getData(**dataParam)

        #p.plot(range(0, len(data)), data, color = line["color"], label = s2)

        #print data
        l = []
        for j in range(1, dataCol):
            l.append(sheet.cell('%s%s'%( get_column_letter(j), i)).value)
        p.plot(range(0, len(data)), data, label = " - ".join(l))

    #p.xlabel(timeLabels, rotation='vertical', size = 'x-small')
    ax = p.gca()
    ax.yaxis.label.set_size('x-small')
    p.xticks(range(len(timeLabels)), timeLabels, rotation='vertical', size = 'x-small', multialignment = 'center')
    #p.xticks(map(add, range(len(data)), [-0.5] * len(data)), timeLabels, rotation='vertical', size = 'x-small', ha = 'right')

    p.legend(loc='best', prop={"size": 'small'}).draw_frame(False)
    p.grid(True)
    p.tight_layout()

    imgName = os.path.join(Settings.tmpPath, "img"  + sheet.title + ".png")

    p.savefig(imgName)
    img = Image(imgName)
    img.anchor(sheet['B%s'%(endRow + 1)], anchortype='oneCell')
    sheet.add_image(img)


def addExcelGraph(sheet, c):
    global dataCol
    dataCol = c + 1
    #print dataCol
    dataParam = {}
    dataParam['sheet'] = sheet

    init(sheet = sheet)

    labels = Reference(sheet, (startRow,  dataCol), (startRow, dataCol + len(timeLabels)))
    chart = LineChart()

    for i in range(startRow + 1, endRow):
        #xvalues = Reference(sheet, (startRow + 1,  dataCol), (startRow + 1, dataCol + len(timeLabels)))
        title = []
        for ii in range(1, dataCol):
            title.append(sheet.cell('%s%s'%(get_column_letter(ii), i)).value)
        values = Reference(sheet, (i,  dataCol), (i, dataCol + len(timeLabels)))
        #print "Reference", (dataCol, startRow + 1), (dataCol + len(timeLabels), startRow + 1)
        series = Series(values, title=" - ".join(title), labels = labels)
        chart.append(series)

    sheet.add_chart(chart)


if __name__ == '__main__':
    try:
        main()
    except (KeyboardInterrupt):
        print "KeyboardInterrupt"

