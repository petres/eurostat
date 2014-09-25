import os, sys
sys.path.insert(1,os.path.join(os.path.dirname(__file__), "..", "lib"))


from openpyxl import Workbook, load_workbook, styles
from openpyxl.cell import get_column_letter
from openpyxl.styles import Border, Style, Side, borders, fills, Fill, Color

#import inspect
#print inspect.getfile(Workbook)

# ADD FOR ADDING TUPLES
from operator import add

import helpers as f

class Writer():

    border = Side(style = borders.BORDER_THIN)
    allBorders = Border(left = border, right = border, top = border, bottom = border)
    dataStyle = Style(border = allBorders)
    labelStyle = Style(border = allBorders, fill = fills.PatternFill(fill_type = fills.FILL_SOLID, fgColor = Color("99ccff"), bgColor = Color("99ccff")))
    flagStyle = Style(border = allBorders, fill = fills.PatternFill(fill_type = fills.FILL_SOLID, fgColor = Color("AAAAAA"), bgColor = Color("AAAAAA")))

    #dataStyle   = None
    #labelStyle  = None
    #flagStyle   = None


    def __init__(self, options):
        self.overwrite = options["overwrite"]
        self.fileName = options["fileName"]
        self.existingSheets = []
        self.opened = False
        #self.sheetCreated = False


    def open(self):
        if self.overwrite == "File":
            wb = Workbook()
        else:
            try:
                wb = load_workbook(self.fileName)
                self.existingSheets = wb.get_sheet_names()
            except Exception as e:
                wb = Workbook()

        self.wb = wb
        self.opened = True


    def changeActiveSheet(self, name):
        if not self.opened:
            self.open()

        #if name == "Sheet":
        #    self.sheetCreated = True

        if name in self.existingSheets:
            if self.overwrite == "Sheet":
                self.wb.remove_sheet(self.wb[name])
                ws = self.wb.create_sheet()
                ws.title = name
            else:
                ws = self.wb[name]
        else:
            ws = self.wb.create_sheet()
            ws.title = name

        self.ws = ws


    def writeHeader(self, options):
        self.write((0, 0), "Name:")
        self.write((0, 1), options["name"])

        self.write((2, 0), "Preset:")
        self.write((2, 1), f.getStringOfPreset(options))

        info = f.getFileInfo(options["name"])
        self.write((3, 0), "Last updated:")
        self.write((3, 1), info["updatedDate"])

        self.write((4, 0), "Extracted on:")
        self.write((4, 1), info["extractedDate"])


    def write(self, coords, value, style = None):
        self.ws.cell('%s%s'%(get_column_letter(coords[1] + 1), coords[0] + 1)).value = value
        if style is not None:
            self.ws.cell('%s%s'%(get_column_letter(coords[1] + 1), coords[0] + 1)).style = style


    def writeTable(self, table, options, sheetName = None, initialOffset = (6, 0)):
        if sheetName is not None:
            changeActiveSheet(sheetName)


        # Row Labels Labels
        rowLabels = []
        tableOffsetRow = 0
        for label in table["structure"]["row"]:
            if label["count"] > 1:
                rowLabels.append(label["name"])
            else:
                self.write((initialOffset[0] + tableOffsetRow, 0), label["name"])
                self.write((initialOffset[0] + tableOffsetRow, 1), label["value"])
                self.write((initialOffset[0] + tableOffsetRow, 2), f.findInDict(label["name"], label["value"]))
                tableOffsetRow += 1

        for crit in table["structure"]["fixed"]:
            self.write((initialOffset[0] + tableOffsetRow, 0), crit)
            self.write((initialOffset[0] + tableOffsetRow, 1), table["structure"]["fixed"][crit])
            self.write((initialOffset[0] + tableOffsetRow, 2), f.findInDict(crit, table["structure"]["fixed"][crit]))
            tableOffsetRow += 1

        tableOffset = (tableOffsetRow + 2, 0)
        offset = map(add, initialOffset, tableOffset)

        self.write(offset, "Data:")
        offset = (offset[0] + 1, offset[1])


        for i, label in enumerate(rowLabels):
            self.write((offset[0], offset[1] + i), label.upper(), self.labelStyle)

        labelOffset = (1, len(rowLabels))

        # Labels
        for i, labels in enumerate(table["labels"]["row"]):
            k = 0
            for j, label in enumerate(labels):
                if table["structure"]["row"][j]["count"] > 1:
                    self.write((offset[0] + i + labelOffset[0], offset[1] + k), label, self.labelStyle)
                    k += 1

        for i, labels in enumerate(table["labels"]["col"]):
            self.write((offset[0], offset[1] + i + labelOffset[1]), " - ".join(labels), self.labelStyle)

        dataOffset = map(add, offset, labelOffset)

        # Data
        for i, line in enumerate(table["data"]):
            for j, entry in enumerate(line):
                style = self.dataStyle

                if entry["flag"] is not None:
                    style = self.flagStyle

                value = entry["value"]
                if value == None:
                    value = options["emptyCellSign"]
                try:
                    self.write((dataOffset[0] + i, dataOffset[1] + j), float(value), style)
                except:
                    self.write((dataOffset[0] + i, dataOffset[1] + j), value, style)

        return (dataOffset[0] + len(table["data"]), dataOffset[1] + len(table["data"][0]))


    def save(self):
        #if not self.sheetCreated:
        #    self.wb.remove_sheet(self.wb["Sheet"])
        self.wb.save(filename = self.fileName)