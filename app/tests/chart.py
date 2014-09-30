import os, sys
sys.path.insert(1, os.path.join(os.path.dirname(__file__), "..", "lib"))


from openpyxl import Workbook
from openpyxl.charts import BarChart, LineChart, Reference, Series
from openpyxl.cell import get_column_letter

wb = Workbook()
ws = wb.active
for i in range(10):
    ws.cell('%s%s'%(get_column_letter(i + 2),  1)).value = i


values = Reference(ws, (1, 2), (1, 11))
title = Reference(ws, (1, 1), (1, 1))
series = Series(values, title=title)
chart = LineChart()
chart.append(series)
ws.add_chart(chart)
wb.save("SampleChart.xlsx")